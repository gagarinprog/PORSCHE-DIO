"""
Regenera data/aggregates.json a partir do Excel original do dashboard
(DASH_BOARD_PORSCHE_.xlsx, aba 'Analise' + aba 'Sanitized').

Uso:
    python scripts/export_data.py caminho/para/DASH_BOARD_PORSCHE_.xlsx
"""

import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

OUT_PATH = Path(__file__).resolve().parent.parent / "data" / "aggregates.json"


def export(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Analise"]

    def cell(r, c):
        return ws.cell(row=r, column=c).value

    by_model = []
    for r in range(4, 10):
        name = cell(r, 1)
        if name:
            by_model.append(
                {
                    "model": name,
                    "units": cell(r, 2),
                    "revenue": round(cell(r, 3), 2),
                    "avg_ticket": round(cell(r, 4), 2),
                }
            )

    status = []
    for r in range(5, 15):
        name = cell(r, 6)
        if name:
            status.append({"status": name, "count": cell(r, 7)})

    payment = []
    for r in range(5, 14):
        name = cell(r, 9)
        if name:
            payment.append({"method": name, "count": cell(r, 10)})

    top_states = []
    for r in range(5, 15):
        name = cell(r, 16)
        if name:
            top_states.append({"state": name, "revenue": cell(r, 17)})

    top_sellers = []
    for r in range(19, 29):
        name = cell(r, 6)
        if name:
            top_sellers.append({"seller": name, "revenue": cell(r, 7)})

    kpis = {
        "total_units": cell(13, 2),
        "total_revenue": cell(14, 2),
        "avg_ticket": round(cell(15, 2), 2),
        "maior_venda": cell(16, 2),
        "km_media": round(cell(17, 2), 1),
        "pct_entregues": cell(18, 2),
    }

    # Receita por ano do modelo (aba Sanitized)
    ws_san = wb["Sanitized"]
    headers = [ws_san.cell(row=1, column=c).value for c in range(1, ws_san.max_column + 1)]
    year_idx = headers.index("YearNum") + 1
    price_idx = headers.index("PriceNum") + 1
    rev_by_year = defaultdict(float)
    for r in range(2, ws_san.max_row + 1):
        y = ws_san.cell(row=r, column=year_idx).value
        p = ws_san.cell(row=r, column=price_idx).value
        if y and p:
            rev_by_year[int(y)] += float(p)
    revenue_by_year = [
        {"year": y, "revenue": round(v, 2)} for y, v in sorted(rev_by_year.items())
    ]

    data = {
        "kpis": kpis,
        "by_model": by_model,
        "status": status,
        "payment": payment,
        "top_states": top_states,
        "top_sellers": top_sellers,
        "revenue_by_year": revenue_by_year,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"OK -> {OUT_PATH}")


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "DASH_BOARD_PORSCHE_.xlsx"
    export(xlsx)
